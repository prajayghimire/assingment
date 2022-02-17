import openpyxl


def rowcount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return (sheet.max_row)


def colcount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return (sheet.max_column)


def readdata(filename, sheetname, rollnum, colnum):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row=rollnum, column=colnum).value


def savefile(filename, sheetname, rollnum, colnum, data):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=rollnum, column=colnum).value = data
    workbook.save(filename)
